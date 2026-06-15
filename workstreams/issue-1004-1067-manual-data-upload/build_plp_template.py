"""Build the Pluto (PLP) manual data-upload template — a NEW workbook modelled on AOPA's
ACTR-3098 Data_Upload_V4.xlsx, with one tab per EC screen (from the EC Screens folder) and the columns
defined per screen. Columns: AOPA-baseline for the stream/tank/well tabs (confirmed from ACTR-3098),
screenshot-derived for contract-accounts + compositions. New/uncertain columns are marked [confirm].
Output: workstreams/issue-1004-1067-manual-data-upload/PLP_Data_Upload_Template_V1_DRAFT.xlsx
This is a VALIDATION DRAFT for the Pluto business to confirm/correct (see VALIDATION-EMAIL.md)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"C:\Projects\ChoongYin_OS\workstreams\issue-1004-1067-manual-data-upload\PLP_Data_Upload_Template_V1_DRAFT.xlsx"
FACILITIES = ["<Pluto facility 1>", "<Pluto facility 2>", "<Scarborough>"]  # [confirm Pluto facility list]

# Each tab: (key columns), (value columns: (name, unit)), note. Comments + ZWP_ACTR_REF appended to every tab.
TABS = {
    "Daily Prod Well Status 1": {
        "screen": "Daily Production Well Status 1 (PLU)  ->  PWEL_DAY_STATUS",
        "keys": [("Facility", ""), ("Well Code", ""), ("Date", "yyyy-mm-dd")],
        "vals": [("On Strm Hrs", "hr"), ("Oil", "Sm3"), ("Cond", "Sm3"), ("Gas", "Sm3"),
                 ("Water", "Sm3"), ("Gas Lift", "Sm3"), ("Avg WH Press [confirm]", "kPa"),
                 ("Avg WH Temp [confirm]", "C"), ("Avg Choke Size [confirm]", "%")],
        "note": "AOPA WellTheor baseline + common PWEL attrs. [confirm] exact Pluto column set; also confirm whether PWEL_DAY_STATUS_2 adds columns.",
    },
    "Daily Gas Stream Status": {
        "screen": "Daily Gas Stream Status  ->  STRM_DAY_STREAM_MEAS_GAS",
        "keys": [("Facility", ""), ("Stream Code", ""), ("Date", "yyyy-mm-dd")],
        "vals": [("Grs Vol", "Sm3"), ("Meas Energy", "GJ")],
        "note": "From AOPA StreamGas (GRS_VOL_GAS, MEAS_ENERGY).",
    },
    "Daily Liquid Stream Status": {
        "screen": "Daily Liquid Stream Status  ->  STRM_DAY_STREAM_MEAS_OIL",
        "keys": [("Facility", ""), ("Stream Code", ""), ("Date", "yyyy-mm-dd")],
        "vals": [("Grs Vol", "Sm3"), ("Oil Spec Gravity", ""), ("BS&W", "%")],
        "note": "From AOPA StreamOil (GRS_VOL_OIL, OIL_SG, BS_W). BS&W is PCT->FRAC converted on load.",
    },
    "Daily Water Stream Status": {
        "screen": "Daily Water Stream Status  ->  STRM_DAY_STREAM_MEAS_WAT",
        "keys": [("Facility", ""), ("Stream Code", ""), ("Date", "yyyy-mm-dd")],
        "vals": [("Grs Vol", "m3"), ("OIW Avg", "ppm"), ("OIW Peak", "ppm"),
                 ("Oil in Water", "ppm"), ("Density", "kg/Sm3")],
        "note": "From AOPA StreamWater (GRS_VOL_WAT + OIW + density).",
    },
    "Daily Electrical Stream Status": {
        "screen": "Daily Electrical Stream Status  ->  STRM_DAY_STREAM_MEAS_ELE",
        "keys": [("Facility", ""), ("Stream Code", ""), ("Date", "yyyy-mm-dd")],
        "vals": [("Power Consumption", "kWh"), ("Available Hrs", "hr")],
        "note": "From AOPA StreamElectrical (POWER_CONSUMPTION, ON_STREAM_HRS).",
    },
    "Daily Tank Status - VCF": {
        "screen": "Daily Tank Status - VCF Calc  ->  TANK_DAY_* (confirm class)",
        "keys": [("Facility", ""), ("Tank Code", ""), ("Date", "yyyy-mm-dd")],
        "vals": [("Closing Liquid", "Sm3"), ("Free Water", "m3"), ("Diluent", "%")],
        "note": "From AOPA DailyTank (CLOSING_GRS_VOL, CLOSING_WATER_VOL, DILUENT_CUT PCT->FRAC). [confirm] Pluto VCF tank class + any VCF-specific input (temp/density/API).",
    },
    "Stream Gas Component Analysis": {
        "screen": "Stream Gas Component Analysis  ->  STRM_COMP_ANALYSIS  [NEW — not in AOPA]",
        "keys": [("Facility", ""), ("Stream Code", ""), ("Date", "yyyy-mm-dd"), ("Sample Date", "yyyy-mm-dd")],
        "vals": [("N2 (Nitrogen)", "mol%"), ("CO2 (Carbon Dioxide)", "mol%"), ("C1 (Methane)", "mol%"),
                 ("C2 (Ethane)", "mol%"), ("C3 (Propane)", "mol%"), ("iC4 (i-Butane)", "mol%"),
                 ("nC4 (n-Butane)", "mol%"), ("iC5 (i-Pentane)", "mol%"), ("nC5 (n-Pentane)", "mol%"),
                 ("C6+ (Hexanes+)", "mol%"), ("GCV [confirm]", "MJ/Sm3"), ("Spec Gravity [confirm]", "")],
        "note": "[NEW — confirm] Component set + grain (mol% vs value), and whether GCV/SG are entered or derived. Header fields (Sample No / Sampling Method / Analysis Status) confirm if needed.",
    },
    "Well Gas Component Analysis": {
        "screen": "Well Gas Component Analysis  ->  well-composition class (confirm)  [NEW — not in AOPA]",
        "keys": [("Facility", ""), ("Well Code", ""), ("Date", "yyyy-mm-dd"), ("Sample Date", "yyyy-mm-dd")],
        "vals": [("N2 (Nitrogen)", "mol%"), ("CO2 (Carbon Dioxide)", "mol%"), ("C1 (Methane)", "mol%"),
                 ("C2 (Ethane)", "mol%"), ("C3 (Propane)", "mol%"), ("iC4 (i-Butane)", "mol%"),
                 ("nC4 (n-Butane)", "mol%"), ("iC5 (i-Pentane)", "mol%"), ("nC5 (n-Pentane)", "mol%"),
                 ("C6+ (Hexanes+)", "mol%")],
        "note": "[NEW — confirm] same component set as stream gas comp; confirm the well-composition EC class name.",
    },
    "Daily Contract Account Status": {
        "screen": "Daily Contract Account Status  ->  SCTR_ACC_DAY_STATUS  [NEW — not in AOPA]",
        "keys": [("Business Unit", ""), ("Contract", ""), ("Contract Account Name", ""), ("Date", "yyyy-mm-dd")],
        "vals": [("Volume Qty", "Sm3"), ("Mass Qty", "t"), ("Energy Qty", "GJ"), ("Extra 1 Qty", ""), ("UOM", "")],
        "note": "Columns read from the live screen (Volume/Mass/Energy/Extra1 Qty + UOM). [confirm] which are user-updatable vs allocated/derived.",
    },
    "Daily Contract Acct - Company": {
        "screen": "Daily Contract Account Result - Company  ->  SCTR_ACC_DAY_CPY_STATUS  [NEW]",
        "keys": [("Business Unit", ""), ("Contract", ""), ("Contract Account Name", ""), ("Company", ""), ("Date", "yyyy-mm-dd")],
        "vals": [("Volume Qty", "Sm3"), ("Mass Qty", "t"), ("Energy Qty", "GJ"), ("Extra 1 Qty", ""), ("UOM", "")],
        "note": "[NEW — confirm] company-split variant of Daily Contract Account; confirm the Company key + updatable cols.",
    },
    "Monthly Contract Account Status": {
        "screen": "Monthly Contract Account Status  ->  SCTR_ACC_MTH_STATUS  [NEW]",
        "keys": [("Business Unit", ""), ("Contract", ""), ("Contract Account Name", ""), ("Month", "yyyy-mm")],
        "vals": [("Volume Qty", "Sm3"), ("Mass Qty", "t"), ("Energy Qty", "GJ"), ("Extra 1 Qty", ""), ("UOM", "")],
        "note": "[NEW — confirm] monthly variant; confirm month-grain date + updatable cols.",
    },
    "Monthly Contract Acct - Company": {
        "screen": "Monthly Contract Account Company Status  ->  SCTR_ACC_MTH_CPY_STATUS  [NEW]",
        "keys": [("Business Unit", ""), ("Contract", ""), ("Contract Account Name", ""), ("Company", ""), ("Month", "yyyy-mm")],
        "vals": [("Volume Qty", "Sm3"), ("Mass Qty", "t"), ("Energy Qty", "GJ"), ("Extra 1 Qty", ""), ("UOM", "")],
        "note": "[NEW — confirm] monthly company-split variant.",
    },
}

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
KEY_FILL = PatternFill("solid", fgColor="C55A11")
UNIT_FILL = PatternFill("solid", fgColor="D9E1F2")
WHITE = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(italic=True, color="C00000")

wb = openpyxl.Workbook()
# Readme
rm = wb.active; rm.title = "Readme"
readme = [
    ["PLP ECaaS Manual Data Upload — Template (V1 DRAFT for validation)"],
    [""],
    ["Issues 1004 + 1067. Modelled on the AOPA ACTR-3098 baseline. ONE tab per EC screen."],
    ["This is a DRAFT for the Pluto business to VALIDATE — see the accompanying validation email."],
    [""],
    ["How to read each tab:"],
    ["  Row 1 = note (which EC screen/class the tab maps to).  Row 2 = column headers."],
    ["  Row 3 = units.  Enter data from Row 4 down."],
    ["  ORANGE headers = KEY columns (identify the row).  BLUE headers = updatable values."],
    ["  Every tab ends with Comments (MANDATORY, non-blank) + ZWP_ACTR_REF (upload reference)."],
    [""],
    ["RULES (per Issues 1004/1067 — confirm in the email):"],
    ["  - Numeric only (blank cell = skip; non-numeric = error, except Comments)."],
    ["  - Reject negative inputs; block locked months; only update rows at status <= Verified."],
    ["  - Comment mandatory & non-blank on every updated row."],
    ["  - Runs as the uploading user's role; LAST_UPDATED_BY = uploader; REV_TEXT = 'Upload File <n>'."],
    ["  - Facility/Code dropdowns to be wired to Pluto assets. [confirm] columns marked [confirm]."],
]
for r in readme:
    rm.append(r)
rm["A1"].font = Font(bold=True, size=14); rm.column_dimensions["A"].width = 100

fac_dv_formula = '"' + ",".join(FACILITIES) + '"'
for tab, cfg in TABS.items():
    ws = wb.create_sheet(tab[:31])
    cols = cfg["keys"] + cfg["vals"] + [("Comments", ""), ("ZWP_ACTR_REF", "")]
    # Row 1 note
    ws.cell(row=1, column=1, value=f"Tab maps to: {cfg['screen']}   |   {cfg['note']}").font = NOTE_FONT
    # Row 2 headers, Row 3 units
    nkey = len(cfg["keys"])
    for ci, (name, unit) in enumerate(cols, start=1):
        h = ws.cell(row=2, column=ci, value=name); h.font = WHITE
        h.fill = KEY_FILL if ci <= nkey else HDR_FILL
        h.alignment = Alignment(horizontal="center", wrap_text=True)
        u = ws.cell(row=3, column=ci, value=(f"[{unit}]" if unit else ""))
        u.fill = UNIT_FILL; u.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, min(22, len(name) + 2))
    ws.freeze_panes = "A4"
    # Facility/Business-Unit dropdown on the first key column (A), rows 4..200
    dv = DataValidation(type="list", formula1=fac_dv_formula, allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"A4:A200")
wb.create_sheet("RevInfo").append(["Date", "Version", "Change", "Who"])
wb["RevInfo"].append(["2026-06-16", "V1 DRAFT", "Initial Pluto template from AOPA baseline + EC Screens; for business validation", "Claude/Choong-Yin"])
wb.save(OUT)
print("WROTE:", OUT)
print("TABS:", ", ".join(wb.sheetnames))
