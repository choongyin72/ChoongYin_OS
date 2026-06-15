"""READ-ONLY: dump the AOPA upload template ACTR-3098 Data_Upload_V4.xlsx — every sheet/tab + its
header row(s) + a sample data row + any data-validation (dropdown) ranges. So we can validate the
Pluto template's tabs/columns. No writes."""
import openpyxl

PATH = r"C:\Projects\Woodside\jiras\Issue-1004 & 1067  - PLP ECaaS Manual data upload template\sources\ACTR-3098 Data_Upload_V4.xlsx"
wb = openpyxl.load_workbook(PATH, data_only=True, read_only=False)
print(f"WORKBOOK: {len(wb.sheetnames)} sheets\n")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n{'='*78}\nTAB: '{sn}'   (dims {ws.dimensions}, max_col={ws.max_column}, max_row={ws.max_row})")
    # find the header row = first row with >=2 non-empty cells; print up to 3 leading rows
    for r in range(1, min(ws.max_row, 4) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 40) + 1)]
        vals = ["" if v is None else str(v).strip()[:22] for v in vals]
        if any(vals):
            print(f"  R{r}: " + " | ".join(v for v in vals if v != "") if any(vals) else "")
    # data-validation dropdowns on this sheet
    try:
        dvs = ws.data_validations.dataValidation
        for dv in dvs[:6]:
            f = (dv.formula1 or "")[:60]
            print(f"  [dropdown] cells={dv.sqref} src={f}")
    except Exception:
        pass
print("\nDONE")
